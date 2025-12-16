#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pattern Worksheet Generator - Final Optimized Version
1. Header removed for Speaking II & III (Space saving)
2. Capitalized Header (Pattern, Evaluation, Remark)
3. Evaluation circles added (Ⓐ Ⓑ Ⓒ)
"""

try:
    from flask import Flask, render_template, request, send_file, jsonify
    import openpyxl
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.lib import colors
    import os
    import glob
    import random
    from datetime import datetime
    from io import BytesIO
    from urllib.parse import unquote
except ImportError as e:
    print(f"CRITICAL ERROR: {e}")
    raise e

app = Flask(__name__)

# --- 경로 설정 ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FOLDER = os.path.join(BASE_DIR, 'databases')
TEMPLATE_DIR = os.path.join(BASE_DIR, 'templates')
FONT_DIR = os.path.join(BASE_DIR, 'fonts')

app.template_folder = TEMPLATE_DIR

# --- 폰트 설정 ---
def setup_korean_font():
    try:
        local_font = os.path.join(FONT_DIR, 'NanumGothic.ttf')
        if os.path.exists(local_font):
            pdfmetrics.registerFont(TTFont('KoreanFont', local_font))
            return 'KoreanFont'
        return 'Helvetica' 
    except:
        return 'Helvetica'

KOREAN_FONT = setup_korean_font()

def load_patterns_from_excel(filename):
    decoded_filename = unquote(filename)
    file_path = os.path.join(DB_FOLDER, decoded_filename)
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"DB 파일을 찾을 수 없습니다: {decoded_filename}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    ws_overview = wb["Pattern Overview"]
    pattern_info = {}
    for row in ws_overview.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            pattern_info[int(row[0])] = {
                'number': int(row[0]),
                'name': str(row[1]),
                'unit': str(row[3]) if len(row) > 3 and row[3] else 'Level A'
            }
            
    ws_detail = wb["Pattern Details"]
    patterns = {}
    for row in ws_detail.iter_rows(min_row=2, values_only=True):
        try:
            p_num = int(row[0])
            section = row[2]
            content = row[4]
            answer = row[5] if len(row) > 5 and row[5] else ""
            
            if p_num not in patterns:
                patterns[p_num] = {
                    'pattern_num': p_num,
                    'pattern_name': pattern_info.get(p_num, {}).get('name', ''),
                    'unit': pattern_info.get(p_num, {}).get('unit', 'Level A'),
                    'speaking1': [], 'speaking2': [], 'unscramble': []
                }
            
            if section == 'Speaking I':
                patterns[p_num]['speaking1'].append(content)
            elif section == 'Speaking II':
                patterns[p_num]['speaking2'].append((content, answer))
            elif section == 'Unscramble':
                scrambled = row[6].strip('()') if row[6] else ""
                patterns[p_num]['unscramble'].append((content, scrambled, answer))
        except:
            continue
            
    return patterns

def distribute_questions(selected_patterns, target_count=5):
    result = {'speaking1': [], 'speaking2': [], 'unscramble': []}
    if not selected_patterns: return result
    
    pattern_count = len(selected_patterns)
    items_per = target_count // pattern_count
    remainder = target_count % pattern_count
    
    for section in ['speaking1', 'speaking2', 'unscramble']:
        for i, p in enumerate(selected_patterns):
            count = items_per + (1 if i < remainder else 0)
            pool = p[section][:]
            random.shuffle(pool)
            result[section].extend(pool[:count])
            
    return result

def create_worksheet_in_memory(pattern_data, selected_patterns, book_title, student_name="", student_date=""):
    buffer = BytesIO()
    
    # 여백 설정 (상하 여백을 조금 줄여서 한 페이지 확보 유리하게 설정)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=10*mm,
        bottomMargin=8*mm,
        leftMargin=15*mm,
        rightMargin=15*mm
    )
    
    story = []
    p_nums = ", ".join([str(p['pattern_num']) for p in selected_patterns])
    clean_book_title = book_title.replace('.xlsx', '')
    
    # 스타일 정의
    title_style = ParagraphStyle('Title', fontSize=12, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=2)
    section_style = ParagraphStyle('Section', fontSize=11, fontName='Helvetica-Bold', spaceBefore=0, spaceAfter=0)
    item_style = ParagraphStyle('Item', fontSize=10, fontName='Helvetica', leftIndent=0, spaceBefore=2, spaceAfter=2)
    item_kr_style = ParagraphStyle('ItemKr', fontSize=10, fontName=KOREAN_FONT, leftIndent=0, spaceBefore=2, spaceAfter=2)
    line_style = ParagraphStyle('Line', fontSize=10, fontName='Helvetica', spaceAfter=0)
    
    # 테이블 내부 텍스트 스타일
    cell_text_style = ParagraphStyle('CellText', fontSize=10, fontName=KOREAN_FONT, leading=12, alignment=TA_LEFT)
    
    # 평가 원문자 스타일 (크게)
    eval_text_style = ParagraphStyle('EvalText', fontSize=12, fontName=KOREAN_FONT, alignment=TA_CENTER, leading=14)

    # === [헤더 영역] ===
    story.append(Paragraph("<b>Weekly Test</b>", title_style))
    story.append(Paragraph(f"<b>{clean_book_title} - Patterns: {p_nums}</b>", title_style))
    
    display_name = f"NAME: {student_name}" if student_name else "NAME: _______________________________"
    display_date = f"DATE: {student_date}" if student_date else "DATE: _____ / _____"
    
    name_date_data = [[
        Paragraph(display_name, ParagraphStyle('Name', fontSize=12, fontName=KOREAN_FONT)), 
        Paragraph(display_date, ParagraphStyle('Date', fontSize=12, fontName=KOREAN_FONT, alignment=TA_RIGHT))
    ]]
    name_date_table = Table(name_date_data, colWidths=[120*mm, 50*mm])
    name_date_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
    ]))
    story.append(name_date_table)
    story.append(Spacer(1, 3*mm))
    
    # === [테이블 생성 함수 개선] ===
    # include_header 옵션 추가: False일 경우 헤더 행을 생성하지 않음
    def create_section_table(data_rows, include_header=True):
        col_widths = [105*mm, 45*mm, 30*mm]
        
        table_data = []
        if include_header:
            # [수정 2] 대문자 적용: Pattern, Evaluation, Remark
            table_data.append(['Pattern', 'Evaluation', 'Remark'])
            
        for row in data_rows:
            table_data.append(row)
            
        t = Table(table_data, colWidths=col_widths, rowHeights=28)
        
        # 기본 스타일 (전체 테두리 및 정렬)
        styles = [
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]
        
        if include_header:
            # 헤더가 있는 경우: 첫 줄 스타일 적용
            styles.append(('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'))
            styles.append(('ALIGN', (0,0), (-1,0), 'CENTER'))
            styles.append(('LINEBELOW', (0,0), (-1,0), 1.5, colors.black))
            content_start = 1
        else:
            # 헤더가 없는 경우: 0번 행부터 내용
            content_start = 0
            
        # 내용 스타일
        styles.extend([
            ('FONTNAME', (0, content_start), (-1,-1), KOREAN_FONT),
            ('ALIGN', (0, content_start), (0,-1), 'LEFT'),   # Pattern 열 왼쪽 정렬
            ('LEFTPADDING', (0, content_start), (0,-1), 8),
            ('ALIGN', (1, content_start), (-1,-1), 'CENTER'), # Eval, Remark 열 가운데 정렬
        ])
        
        t.setStyle(TableStyle(styles))
        return t
    
    # [수정 3] 평가 기호 적용
    eval_str = "Ⓐ   Ⓑ   Ⓒ" 
    # 만약 폰트 호환 문제로 안 보이면 일반 텍스트 (A) (B) (C)로 대체 가능하나, 
    # NanumGothic은 대부분 지원합니다. Paragraph로 감싸서 스타일 적용.
    eval_para = Paragraph(eval_str, eval_text_style)

    # === [Speaking I] - 헤더 포함 ===
    story.append(Paragraph("<b>◈ Speaking I - Say the meaning</b>", section_style))
    story.append(Spacer(1, 2*mm))
    
    rows_s1 = []
    for idx, question in enumerate(pattern_data['speaking1'][:5], 1):
        text_para = Paragraph(f"{idx}. {question}", cell_text_style)
        rows_s1.append([text_para, eval_para, ""])
        
    # include_header=True (기본값)
    story.append(create_section_table(rows_s1, include_header=True))
    story.append(Spacer(1, 3*mm))
    
    # === [Speaking II] - 헤더 제거 ===
    story.append(Paragraph("<b>◈ Speaking II - Say in English</b>", section_style))
    story.append(Spacer(1, 2*mm))
    
    rows_s2 = []
    for idx, (korean, answer) in enumerate(pattern_data['speaking2'][:5], 1):
        text_para = Paragraph(f"{idx+5}. {korean}", cell_text_style)
        rows_s2.append([text_para, eval_para, ""])
        
    # [수정 1] include_header=False
    story.append(create_section_table(rows_s2, include_header=False))
    story.append(Spacer(1, 3*mm))
    
    # === [Speaking III] - 헤더 제거 ===
    story.append(Paragraph("<b>◈ Speaking III - Talk with your teacher</b>", section_style))
    story.append(Spacer(1, 2*mm))
    
    rows_s3 = []
    for i in range(5):
        pat_num_str = f"Pattern {i+1}"
        if i < len(selected_patterns):
            pat_num_str = f"Pattern {selected_patterns[i]['pattern_num']}"
            
        text_para = Paragraph(f"{i+11}. {pat_num_str}", cell_text_style)
        rows_s3.append([text_para, eval_para, ""])
        
    # [수정 1] include_header=False
    story.append(create_section_table(rows_s3, include_header=False))
    story.append(Spacer(1, 3*mm))
    
    # === [Unscramble] ===
    story.append(Paragraph("<b>◈ Unscramble</b>", section_style))
    story.append(Spacer(1, 2*mm))
    for idx, (korean, words, answer) in enumerate(pattern_data['unscramble'][:5], 1):
        story.append(Paragraph(f"{idx}. {korean} ({words})", item_kr_style))
        story.append(Spacer(1, 6*mm))  # 간격 미세 조정
        story.append(Paragraph("_" * 85, line_style))
        story.append(Spacer(1, 2*mm))
    
    story.append(Spacer(1, 4*mm))
    
    # === [Footer] ===
    footer_data = [[
        Paragraph("<b>GRADE:</b>", ParagraphStyle('Footer', fontSize=12, fontName='Helvetica-Bold')),
        "",
        Paragraph("<b>REMARK:</b>", ParagraphStyle('Footer', fontSize=12, fontName='Helvetica-Bold'))
    ]]
    footer_table = Table(footer_data, colWidths=[40*mm, 40*mm, 90*mm])
    footer_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (2, 0), (2, 0), 'LEFT'),
    ]))
    story.append(footer_table)
    
    # === PAGE 2 (Answer Key) ===
    story.append(PageBreak())
    story.append(Paragraph("<b>Teacher's Guide (Answer Key)</b>", title_style))
    story.append(Paragraph(f"<b>{clean_book_title} - Patterns: {p_nums}</b>", title_style))
    story.append(Spacer(1, 10*mm))
    
    story.append(Paragraph("<b>◈ Speaking II Answers</b>", section_style))
    story.append(Spacer(1, 3*mm))
    for idx, (korean, answer) in enumerate(pattern_data['speaking2'][:5], 1):
        story.append(Paragraph(f"<b>{idx}.</b> {answer}", item_style))
    story.append(Spacer(1, 10*mm))
    
    story.append(Paragraph("<b>◈ Unscramble Answers</b>", section_style))
    story.append(Spacer(1, 3*mm))
    for idx, (korean, words, answer) in enumerate(pattern_data['unscramble'][:5], 1):
        story.append(Paragraph(f"<b>{idx}.</b> {answer}", item_style))

    doc.build(story)
    buffer.seek(0)
    return buffer

@app.route('/')
def index():
    try:
        files = glob.glob(os.path.join(DB_FOLDER, "*.xlsx"))
        books = sorted([os.path.basename(f) for f in files])
        return render_template('index.html', books=books)
    except Exception as e:
        return f"<h3>Error: {str(e)}</h3><p>Check logs for details.</p>"

@app.route('/get_patterns/<filename>')
def get_patterns(filename):
    try:
        decoded_filename = unquote(filename)
        patterns = load_patterns_from_excel(decoded_filename)
        pattern_list = []
        for p_num in sorted(patterns.keys()):
            pattern_list.append({
                'number': p_num,
                'name': patterns[p_num]['pattern_name'],
                'unit': patterns[p_num]['unit']
            })
        return jsonify({'success': True, 'patterns': pattern_list})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/generate', methods=['POST'])
def generate():
    try:
        data = request.json
        book_filename = data.get('book')
        selected_nums = data.get('patterns', [])
        student_name = data.get('name', '')
        student_date = data.get('date', '')
        
        if not book_filename or not selected_nums:
            return jsonify({'error': 'Book or Patterns missing'}), 400
            
        all_patterns = load_patterns_from_excel(book_filename)
        selected_data = []
        for num in selected_nums:
            if int(num) in all_patterns:
                selected_data.append(all_patterns[int(num)])
                
        final_questions = distribute_questions(selected_data)
        
        pdf_buffer = create_worksheet_in_memory(
            final_questions, selected_data, book_title=book_filename, 
            student_name=student_name, student_date=student_date
        )
        
        filename = f"Worksheet_{datetime.now().strftime('%m%d_%H%M%S')}.pdf"
        
        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)